# -*- coding: utf-8 -*-
import io
import os
import re
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List
from datetime import datetime
import re
import time
import tempfile
from zipfile import BadZipFile
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from PIL import Image
from openpyxl import Workbook, load_workbook

# ---------- Path & filesystem ----------

def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)

def sanitize_for_name(s: str) -> str:
    s = re.sub(r"[\\/]+", "_", s)
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s)
    return s.strip("_")

def atomic_write_text(path: Path, text: str, encoding: str = "utf-8") -> None:
    ensure_dir(path.parent)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with open(tmp, "w", encoding=encoding) as f:
        f.write(text)
    os.replace(tmp, path)

def save_checkpoint_json(base_dir: Path, pdf_rel: str, suffix: str, obj: Dict[str, Any]) -> Path:
    """
    Salva un JSON di checkpoint per file/batch.
    """
    ensure_dir(base_dir)
    base = sanitize_for_name(pdf_rel)
    out = base_dir / f"{base}{suffix}.json"
    atomic_write_text(out, json.dumps(obj, ensure_ascii=False, indent=2))
    return out

# ---------- PDF discovery ----------

def iter_pdfs(root: Path) -> List[Path]:
    return [p for p in root.rglob("*.pdf") if p.is_file()]

# ---------- Imaging ----------

def encode_png_base64(pil_img: Image.Image) -> str:
    buf = io.BytesIO()
    pil_img.save(buf, format="PNG")
    import base64
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("utf-8")

# ---------- Excel (4 colonne) ----------

HEADERS_4C = ["file", "data", "modello", "quantità"]

def append_4cols_xlsx(xlsx_path: Path, rows_4: list[dict]) -> None:
    """
    Appende righe al file XLSX finale con colonne fisse:
    [file, date, model, quantity]

    - Se il file esiste ma non è un XLSX valido, lo mette in backup e ne crea uno nuovo.
    - Scrittura atomica con retry (tmp → replace) per evitare corruzioni e gestire lock su Windows.
    - Scarta righe con quantity non valida.
    """
    if not rows_4:
        return

    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    # --- helper: coerce/validate row ---
    def _coerce_row(d: dict) -> tuple[str, str, str, int] | None:
        f = str(d.get("file", "")).strip()
        dt = str(d.get("date") or d.get("order_date") or "").strip()
        m  = str(d.get("model", "")).strip()

        qv = d.get("quantity")
        q = None
        if isinstance(qv, int):
            q = qv
        elif isinstance(qv, str):
            m_num = re.search(r"\d+", qv.replace(" ", ""))
            q = int(m_num.group()) if m_num else None
        else:
            try:
                q = int(qv) if qv is not None else None
            except Exception:
                q = None

        if q is None or q < 1 or q > 2_000_000:
            return None

        return (f, dt, m, q)

    # --- carica o crea workbook ---
    if xlsx_path.exists() and xlsx_path.stat().st_size > 0:
        try:
            wb = load_workbook(xlsx_path)
        except (InvalidFileException, BadZipFile, KeyError):
            # file esistente ma non valido → backup e nuovo workbook
            backup = xlsx_path.with_suffix(
                xlsx_path.suffix + f".corrupted.{time.strftime('%Y%m%d-%H%M%S')}.bak"
            )
            try:
                xlsx_path.replace(backup)
                print(f"[utils] ⚠️ XLSX corrotto: spostato in {backup.name}")
            except Exception:
                pass
            wb = Workbook()
    else:
        wb = Workbook()

    ws = wb.active
    if ws.title == "Sheet":
        ws.title = "data"

    header = ["file", "date", "model", "quantity"]
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(header)
    else:
        current_header = [ws.cell(1, i+1).value for i in range(len(header))]
        if current_header != header:
            if "data" in wb.sheetnames:
                ws = wb["data"]
                if ws.max_row == 1 and ws["A1"].value is None:
                    ws.append(header)
            else:
                ws = wb.create_sheet("data")
                ws.append(header)

    # --- append rows ---
    appended = 0
    for r in rows_4:
        row = _coerce_row(r)
        if row is None:
            continue
        ws.append(row)
        appended += 1

    # --- salvataggio atomico con retry ---
    fd, tmp_path_str = tempfile.mkstemp(prefix="xlsx_tmp_", suffix=".xlsx", dir=str(xlsx_path.parent))
    os.close(fd)
    tmp = Path(tmp_path_str)
    try:
        wb.save(tmp)
    finally:
        wb.close()

    # Prova a sostituire il target con retry
    max_tries = 6  # ~6 secondi di retry
    for i in range(max_tries):
        try:
            os.replace(tmp, xlsx_path)
            print(f"[utils] 💾 XLSX aggiornato: {xlsx_path.name} (righe aggiunte: {appended})")
            return
        except PermissionError as e:
            # file bloccato da Excel o altro processo
            time.sleep(1.0)
        except Exception as e:
            # altro errore di FS: breve retry
            time.sleep(0.5)

    # Se ancora bloccato, salva con nome alternativo per non perdere dati
    alt = xlsx_path.with_name(
        f"{xlsx_path.stem}.inuse-{time.strftime('%Y%m%d-%H%M%S')}{xlsx_path.suffix}"
    )
    try:
        os.replace(tmp, alt)
        print(f"[utils] ⚠️ File in uso, salvato come: {alt.name} (unisci quando libero)")
    except Exception as e:
        # estremo fallback: lasciare il tmp in chiaro
        print(f"[utils] ❌ Impossibile scrivere {xlsx_path.name} e {alt.name}: lascio il tmp {tmp.name}")

# ---------- OpenAI Responses helpers ----------

def safe_find_json_block(text: str) -> str:
    # rimuovo eventuali fence
    text = re.sub(r"```json|```", "", text, flags=re.IGNORECASE)
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("Nessun blocco JSON trovato nell'output del modello.")
    return text[start:end+1]

def extract_text_from_response(resp) -> str:
    """
    Supporta l'SDK 'responses' (OpenAI Python >= 1.0).
    Prova vari rami noti per ricostruire l'output testuale.
    """
    txt = getattr(resp, "output_text", None)
    if isinstance(txt, str) and txt.strip():
        return txt

    # struttura alternativa
    if hasattr(resp, "output") and isinstance(resp.output, list):
        chunks: List[str] = []
        for item in resp.output:
            if getattr(item, "type", None) == "output_text":
                t = getattr(item, "text", None)
                val = getattr(t, "value", None)
                if isinstance(val, str):
                    chunks.append(val)
        if chunks:
            return "".join(chunks)

    # fallback a model_dump (pydantic)
    try:
        d = resp.model_dump()
        if isinstance(d, dict):
            if isinstance(d.get("output_text"), str):
                return d["output_text"]
            chunks = []
            for it in d.get("output", []):
                if isinstance(it, dict) and it.get("type") == "output_text":
                    tv = ((it.get("text") or {}) if isinstance(it.get("text"), dict) else {}).get("value")
                    if isinstance(tv, str):
                        chunks.append(tv)
            if chunks:
                return "".join(chunks)
    except Exception:
        pass

    raise RuntimeError("Impossibile ricostruire il testo dalla risposta dell'SDK OpenAI.")

def log(msg: str) -> None:
    """Stampa con timestamp locale [HH:MM:SS]."""
    try:
        now = datetime.now().strftime("%H:%M:%S")
    except Exception:
        now = "--:--:--"
    print(f"[{now}] {msg}")

def preview_rows(rows, n: int = 5) -> str:
    """Restituisce una piccola anteprima delle righe (file, data, modello, quantità)."""
    if not rows:
        return "(nessuna riga)"
    out = []
    for r in rows[:n]:
        f = r.get("file", "")
        d = r.get("order_date", r.get("data", ""))
        m = r.get("model", r.get("modello", ""))
        q = r.get("quantity", r.get("quantità", ""))
        out.append(f"- {f} | date={d} | model={m} | qty={q}")
    more = "" if len(rows) <= n else f" (+{len(rows)-n} righe)"
    return "\n".join(out) + more
